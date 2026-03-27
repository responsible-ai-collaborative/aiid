from __future__ import annotations

from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _col_group(column: str) -> str:
    identity = {"Incident ID", "date", "year", "title", "description", "deployer", "developer", "harmed"}
    coverage = {"Data Sources", "report_count"}
    mit = {"Risk Domain", "Risk Subdomain", "Responsible Entity", "Intent", "Timing"}
    gmf = {"AI Goal", "AI Technology", "Technical Failure"}

    if column in identity:
        return "identity"
    if column in coverage:
        return "coverage"
    if column in mit:
        return "mit"
    if column in gmf:
        return "gmf"
    return "cset"


def _write_master_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Master Dataset")
    cols = df.columns.tolist()
    n = len(cols)

    hdr_fills = {
        "identity": PatternFill("solid", fgColor="1F3864"),
        "coverage": PatternFill("solid", fgColor="2E4057"),
        "mit": PatternFill("solid", fgColor="1A6B3C"),
        "gmf": PatternFill("solid", fgColor="7B3F00"),
        "cset": PatternFill("solid", fgColor="4A235A"),
    }
    group_hex = {k: v.fgColor.rgb for k, v in hdr_fills.items()}

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    body_font = Font(name="Arial", size=9)
    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    thin_side = Side(style="thin", color="D0D0D0")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    group_labels = {
        "identity": "INCIDENT IDENTITY (incidents.csv - 100% complete)",
        "coverage": "COVERAGE (derived)",
        "mit": "RISK CLASSIFICATION - MIT",
        "gmf": "TECHNICAL ANALYSIS - GMF",
        "cset": "POLICY DETAIL - CSETv1",
    }

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n)
    title_cell = ws.cell(
        row=1,
        column=1,
        value=f"AI Incident Database - Master Dataset | {len(df):,} incidents | {n} columns",
    )
    title_cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor="0D1B2A")
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    current_group, group_start = None, 1
    for idx, col in enumerate(cols, 1):
        group = _col_group(col)
        if group != current_group:
            if current_group is not None:
                ws.merge_cells(start_row=2, start_column=group_start, end_row=2, end_column=idx - 1)
                cell = ws.cell(row=2, column=group_start)
                cell.value = group_labels.get(current_group, "")
                cell.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=group_hex[current_group])
                cell.alignment = Alignment(horizontal="left", vertical="center")
            current_group, group_start = group, idx
    ws.merge_cells(start_row=2, start_column=group_start, end_row=2, end_column=n)
    cell = ws.cell(row=2, column=group_start)
    cell.value = group_labels.get(current_group, "")
    cell.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor=group_hex[current_group])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    for idx, col in enumerate(cols, 1):
        cell = ws.cell(row=3, column=idx, value=col)
        cell.font = hdr_font
        cell.fill = hdr_fills[_col_group(col)]
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[3].height = 36

    for row_idx, row in enumerate(df.itertuples(index=False), start=4):
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            elif hasattr(value, "strftime"):
                value = value.strftime("%Y-%m-%d")
            elif isinstance(value, float) and value == int(value):
                value = int(value)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=False)

    widths = {
        "Incident ID": 11,
        "date": 13,
        "year": 7,
        "title": 40,
        "description": 50,
        "deployer": 25,
        "developer": 25,
        "harmed": 25,
        "Data Sources": 22,
        "report_count": 10,
        "Risk Domain": 30,
        "Risk Subdomain": 42,
        "Responsible Entity": 16,
        "Intent": 14,
        "Timing": 16,
        "AI Goal": 35,
        "AI Technology": 30,
        "Technical Failure": 35,
    }
    for idx, col in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(col, 18)

    ws.freeze_panes = ws.cell(row=4, column=3)
    ws.auto_filter.ref = f"A3:{get_column_letter(n)}3"


def _write_dictionary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Data Dictionary")
    ws.merge_cells("A1:E1")
    title = ws.cell(row=1, column=1, value="Data Dictionary - AI Incident Database Master Dataset")
    title.font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0D1B2A")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 26

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    for idx, header in enumerate(["Column", "Group", "Source", "Fill Rate", "Description"], 1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

    descriptions = {
        "Incident ID": "Primary join key. Unique per incident.",
        "date": "Date harm occurred (editor-resolved).",
        "year": "Year derived from date - best for time series.",
        "title": "Short editor-written title of the incident.",
        "description": "One to three sentence summary of what happened.",
        "deployer": "Who deployed the AI. Cleaned from JSON slug format.",
        "developer": "Who built the AI system.",
        "harmed": "Who was harmed or nearly harmed.",
        "Data Sources": "Which taxonomies classified this incident: MIT | GMF | CSETv1.",
        "report_count": "Number of linked news articles.",
        "Risk Domain": "High-level risk category.",
        "Risk Subdomain": "Granular sub-category nested under Risk Domain.",
        "Responsible Entity": "Who caused the risk: AI / Human / Other.",
        "Intent": "Intentional vs Unintentional vs Other.",
        "Timing": "Pre-deployment vs Post-deployment.",
        "AI Goal": "What the AI was trying to do.",
        "AI Technology": "ML/AI technique used.",
        "Technical Failure": "What technically failed.",
        "Harm Domain": "Whether harm occurred in a recognized domain.",
        "Tangible Harm": "Level of tangible harm.",
        "AI Harm Level": "AI contribution to harm severity.",
        "Rights Violation": "Whether a legal or human rights violation occurred.",
        "Lives Lost": "Fatality count.",
        "Injuries": "Injury count.",
        "Sector of Deployment": "Industry sector (ISIC classification).",
        "Location Region": "World region.",
        "Country Code": "ISO 2-letter country code.",
        "Intentional Harm": "Whether harm was intentionally designed into the system.",
        "Autonomy Level": "Autonomy level of the system.",
    }

    group_colors = {
        "Identity": "1F3864",
        "Coverage": "2E4057",
        "MIT": "1A6B3C",
        "GMF": "7B3F00",
        "CSETv1": "4A235A",
    }
    source_map = {
        "identity": "incidents.csv",
        "coverage": "Derived",
        "mit": "MIT",
        "gmf": "GMF",
        "cset": "CSETv1",
    }

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    body_font = Font(name="Arial", size=9)

    for row_idx, col in enumerate(df.columns, start=3):
        group = _col_group(col)
        group_label = group.upper() if group not in ("identity", "coverage") else group.capitalize()
        fill_pct = f"{df[col].notna().mean()*100:.0f}%"
        desc = descriptions.get(col, "-")
        bg = alt_fill if row_idx % 2 == 0 else white_fill

        for col_idx, value in enumerate([col, group_label, source_map.get(group, "-"), fill_pct, desc], 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = bg
            cell.border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if col_idx == 2:
                cell.font = Font(name="Arial", size=9, bold=True, color=group_colors.get(group_label, "000000"))
            else:
                cell.font = body_font

    for col_letter, width in zip(["A", "B", "C", "D", "E"], [28, 12, 14, 10, 72]):
        ws.column_dimensions[col_letter].width = width
    ws.freeze_panes = "A3"


def _write_coverage_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Coverage Map")
    ws.merge_cells("A1:D1")
    title = ws.cell(row=1, column=1, value="Coverage Map - What you can analyze at each taxonomy level")
    title.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    title.fill = PatternFill("solid", fgColor="0D1B2A")
    title.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    for idx, header in enumerate(["Data Sources", "Incidents", "% of Total", "What you can analyze"], 1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.font = header_font
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

    analysis_map = {
        "MIT": "Risk domain trends, intent, timing, entity - broad lens",
        "MIT | GMF": "MIT analysis plus technical failures and AI goals",
        "MIT | GMF | CSETv1": "Full picture: risk, technical, policy, sector, geography, harm",
        "MIT | CSETv1": "Risk plus policy: sector, lives lost, location, harm level, rights",
        "None": "Title, description, deployer, developer, harmed, report count",
    }

    total = len(df)
    breakdown = df["Data Sources"].value_counts().reset_index()
    breakdown.columns = ["Data Sources", "Count"]

    alt_fill = PatternFill("solid", fgColor="F5F5F5")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    body_font = Font(name="Arial", size=9)

    for row_idx, row in enumerate(breakdown.itertuples(index=False), start=3):
        bg = alt_fill if row_idx % 2 == 0 else white_fill
        pct = f"{row.Count / total * 100:.1f}%"
        values = [row._0, row.Count, pct, analysis_map.get(row._0, "-")]
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = body_font
            cell.fill = bg
            cell.border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    for col_idx, value in enumerate(["TOTAL", total, "100%", ""], 1):
        cell = ws.cell(row=len(breakdown) + 3, column=col_idx, value=value)
        cell.font = Font(name="Arial", bold=True, size=9)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")
        cell.border = Border(
            left=Side(style="thin", color="D0D0D0"),
            right=Side(style="thin", color="D0D0D0"),
            top=Side(style="thin", color="D0D0D0"),
            bottom=Side(style="thin", color="D0D0D0"),
        )

    for col_letter, width in zip(["A", "B", "C", "D"], [24, 14, 12, 68]):
        ws.column_dimensions[col_letter].width = width


def export_excel(master: pd.DataFrame, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _write_master_sheet(wb, master)
    _write_dictionary_sheet(wb, master)
    _write_coverage_sheet(wb, master)

    wb.save(output_path)
